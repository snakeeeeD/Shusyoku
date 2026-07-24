# -*- coding: utf-8 -*-
"""cards_edit.xlsx -> cards.json （Excelの編集をゲームデータに反映）
安全のため cards.json を .bak にバックアップしてから書き出す。
"""
import json
import shutil
import sys
from openpyxl import load_workbook
from cards_common import CARDS_JSON, CARDS_XLSX, COLS, row_to_card

wb = load_workbook(CARDS_XLSX, data_only=True)
ws = wb["Cards"]

header = [ws.cell(1, c).value for c in range(1, len(COLS) + 1)]
if header != COLS:
    print("!! 見出し行が想定と違います。列を変更しないでください。")
    print("   期待:", COLS)
    print("   実際:", header)
    sys.exit(1)

cards = []
errors = []
seen = set()
for r in range(2, ws.max_row + 1):
    row = {name: ws.cell(r, c).value for c, name in enumerate(COLS, 1)}
    if not row.get("id"):
        continue  # 空行はスキップ
    try:
        cd = row_to_card(row)
    except Exception as e:
        errors.append(f"  行{r}: {e}")
        continue
    if cd["id"] in seen:
        errors.append(f"  行{r}: id重複 '{cd['id']}'")
        continue
    seen.add(cd["id"])
    # 最低限の検証
    if not cd.get("name") or not cd.get("type") or not cd.get("rangeType"):
        errors.append(f"  行{r} ({cd['id']}): name/type/rangeType が空")
    cards.append(cd)

if errors:
    print("!! エラーがあるため書き出しを中止しました:")
    print("\n".join(errors))
    sys.exit(1)

# バックアップ
shutil.copyfile(CARDS_JSON, CARDS_JSON + ".bak")

with open(CARDS_JSON, "w", encoding="utf-8") as f:
    json.dump({"cards": cards}, f, ensure_ascii=False, indent=2)

# 検証：書き出した JSON を読み直せるか
with open(CARDS_JSON, encoding="utf-8") as f:
    json.load(f)

print(f"OK: {len(cards)}枚を cards.json に書き出しました（.bak にバックアップ済み）")
