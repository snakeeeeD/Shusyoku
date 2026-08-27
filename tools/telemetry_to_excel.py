#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
telemetry_to_excel.py

ゲームのテレメトリJSONL（logs/session_*.jsonl）を読み込み、1つのExcelブックに集計する。
- 生ログ（JSONL）は絶対に編集しない。読むだけ。出力は毎回作り直せる使い捨て。
- ラン単位（1行=1ラン）＋全体統計サマリ＋各詳細シートを生成する。

使い方:
    python tools/telemetry_to_excel.py [ログのディレクトリ ...] [-o 出力.xlsx]

例:
    python tools/telemetry_to_excel.py                     # 既定の場所を自動探索
    python tools/telemetry_to_excel.py logs -o report.xlsx
    python tools/telemetry_to_excel.py x64/Debug/logs x64/Release/logs
"""
import sys, os, json, glob, argparse
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="2F3B52")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUB_FONT = Font(name=FONT, bold=True, size=11, color="2F3B52")
CELL_FONT = Font(name=FONT)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------- 名前解決（日本語）
CARD_NAME, RELIC_NAME, ENEMY_NAME = {}, {}, {}
CARD_RARITY, CARD_TYPE = {}, {}      # id -> rarity / type
RELIC_RARITY = {}                    # id -> rarity
CARD_ELIGIBLE = []                   # 報酬対象カード（starter/generated除く）id順

# 列見出しの日本語ラベル（表示用。行データのキーは英語のまま）
HEADER_JA = {
    "session": "セッション", "run": "ラン", "t": "時刻",
    "result": "結果", "reached_layer": "到達層", "end_node(steps)": "終了地点(残歩数)",
    "hpLeft": "残HP", "maxHp": "最大HP", "cause": "死因id", "cause_ja": "死因",
    "sec": "プレイ秒", "final_deckSize": "最終デッキ枚数", "gold": "所持金",
    "battles": "戦闘数", "wins": "勝利", "losses": "敗北",
    "cards_picked": "取得カード数", "cards_skipped": "スキップ数",
    "relics_got": "レリック取得", "materials": "素材取得",
    "dmg_taken": "被ダメ合計", "avg_turns": "平均ターン",
    "final_deck": "最終デッキ(id)", "final_deck_ja": "最終デッキ",
    "final_relics": "最終レリック(id)", "final_relics_ja": "最終レリック",
    "layer": "層", "node": "地点(歩数)", "enc": "エンカ", "cat": "種別",
    "tier": "段階", "count": "敵数", "foes": "敵(id)", "foes_ja": "敵",
    "turns": "ターン", "hpStart": "開始HP",
    "chosen": "選択(id)", "chosen_ja": "選択", "offered": "提示(id)", "offered_ja": "提示",
    "id": "id", "n": "個数", "rare": "レア素材",
    "event": "イベント", "choice": "選択肢", "kind": "種類", "price": "価格",
    "name": "名前", "rarity": "レア度", "type": "種類", "marker": "マーカー",
    "scene": "画面", "hp": "HP",
    "detail": "内容(id)", "detail_ja": "内容", "extra": "追加(id)", "extra_ja": "追加",
    "weight": "出現重み", "category": "カテゴリ",
    "get_count": "取得回数", "offered_ct": "提示回数", "picked_ct": "選択回数",
    "pick_rate": "採用率", "appeared": "出現",
}


def _load_json(path, key):
    try:
        d = json.load(open(path, encoding="utf-8-sig"))
        return d if isinstance(d, list) else d.get(key, d)
    except (OSError, json.JSONDecodeError):
        return None


def load_name_maps(script_dir):
    """cards.json / relics.json（name有）と enemy_names_ja.json を読み、id→日本語名を作る。"""
    data_dirs = ["GameEngine/Assets/Data", "Assets/Data",
                 os.path.join(script_dir, "..", "GameEngine", "Assets", "Data")]
    for dd in data_dirs:
        cards = _load_json(os.path.join(dd, "cards.json"), "cards")
        if cards:
            for c in cards:
                if isinstance(c, dict) and c.get("id"):
                    cid = c["id"]
                    CARD_NAME[cid] = c.get("name", cid)
                    CARD_RARITY[cid] = c.get("rarity", "")
                    CARD_TYPE[cid] = c.get("type", "")
                    if not (c.get("starter") or c.get("generated")):
                        CARD_ELIGIBLE.append(cid)
        relics = _load_json(os.path.join(dd, "relics.json"), "relics")
        if relics:
            for r in relics:
                if isinstance(r, dict) and r.get("id"):
                    RELIC_NAME[r["id"]] = r.get("name", r["id"])
                    RELIC_RARITY[r["id"]] = r.get("rarity", "")
        if CARD_NAME and RELIC_NAME:
            break
    em = _load_json(os.path.join(script_dir, "enemy_names_ja.json"), "enemies")
    if isinstance(em, dict):
        for k, v in em.items():
            if not k.startswith("_"):
                ENEMY_NAME[k] = v


def card_ja(cid):
    cid = (cid or "").strip()
    if not cid:
        return ""
    if cid == "(skip)":
        return "(スキップ)"
    if cid.startswith("CRAFT:"):
        return "合成カード"
    if cid.endswith("+"):
        return card_ja(cid[:-1]) + "+"
    return CARD_NAME.get(cid, cid)


def enemy_ja(eid):
    eid = (eid or "").strip()
    return ENEMY_NAME.get(eid, eid) if eid else ""


def relic_ja(rid):
    rid = (rid or "").strip()
    return RELIC_NAME.get(rid, rid) if rid else ""


def join_ja(csv, fn):
    return ",".join(fn(x) for x in str(csv or "").split(",") if x.strip())


# ---------------------------------------------------------------- ログ読み込み
def default_dirs():
    cands = ["logs", "x64/Debug/logs", "x64/Release/logs",
             "GameEngine/logs", "GameEngine/x64/Debug/logs"]
    return [d for d in cands if os.path.isdir(d)]


def load_events(dirs):
    events = []
    files = []
    for d in dirs:
        files += glob.glob(os.path.join(d, "**", "session_*.jsonl"), recursive=True)
        files += glob.glob(os.path.join(d, "session_*.jsonl"))
    for path in sorted(set(files)):
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        e = json.loads(line)
                    except json.JSONDecodeError:
                        continue          # 壊れた行（クラッシュ直後の書きかけ等）は飛ばす
                    e["_file"] = os.path.basename(path)
                    events.append(e)
        except OSError:
            pass
    return events, sorted(set(files))


# ---------------------------------------------------------------- ラン単位に整理
def group_runs(events):
    """(session, run) ごとにイベントをまとめる。時刻順は入力順を維持。"""
    runs = defaultdict(list)
    for e in events:
        key = (e.get("session", ""), e.get("run", 0))
        runs[key].append(e)
    return runs


def num(v, d=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def build_run_rows(runs):
    rows = []
    for (session, run), evs in runs.items():
        end = next((e for e in reversed(evs) if e.get("event") == "run_end"), None)
        start = next((e for e in evs if e.get("event") == "run_start"), None)
        battle_ends = [e for e in evs if e.get("event") == "battle_end"]
        # run_end も battle も無い群（起動直後の空ラン等）はランとして扱わない
        if end is None and not battle_ends:
            continue

        picks = [e for e in evs if e.get("event") == "card_pick"]
        picked = [e for e in picks if e.get("chosen") != "(skip)"]
        skipped = [e for e in picks if e.get("chosen") == "(skip)"]
        dmg_taken = sum(max(0, num(b.get("hpStart")) - num(b.get("hpLeft"))) for b in battle_ends)
        turns = [num(b.get("turns")) for b in battle_ends]

        result = end.get("result") if end else "incomplete"
        rows.append({
            "session": session,
            "run": run,
            "result": result,
            "reached_layer": (end or {}).get("layer", ""),
            "end_node(steps)": (end or {}).get("node", ""),
            "hpLeft": (end or {}).get("hpLeft", ""),
            "maxHp": (end or {}).get("maxHp", ""),
            "cause": (end or {}).get("cause", ""),
            "sec": (end or {}).get("sec", ""),
            "final_deckSize": (end or {}).get("deckSize", ""),
            "gold": (end or {}).get("gold", ""),
            "battles": len(battle_ends),
            "wins": sum(1 for b in battle_ends if b.get("result") == "win"),
            "losses": sum(1 for b in battle_ends if b.get("result") == "lose"),
            "cards_picked": len(picked),
            "cards_skipped": len(skipped),
            "relics_got": sum(1 for e in evs if e.get("event") == "relic_get"),
            "materials": sum(1 for e in evs if e.get("event") == "material_drop"),
            "dmg_taken": int(dmg_taken),
            "avg_turns": round(sum(turns) / len(turns), 2) if turns else "",
            "cause_ja": enemy_ja((end or {}).get("cause", "")),
            "final_deck": (end or {}).get("deck", ""),
            "final_deck_ja": join_ja((end or {}).get("deck", ""), card_ja),
            "final_relics": (end or {}).get("relics", ""),
            "final_relics_ja": join_ja((end or {}).get("relics", ""), relic_ja),
        })
    # 新しいラン（sessionの日時文字列）を上に
    rows.sort(key=lambda r: (str(r["session"]), num(r["run"])), reverse=True)
    return rows


# ---------------------------------------------------------------- 詳細行の抽出
def combats(runs):
    out = []
    for (session, run), evs in runs.items():
        last_start = None
        for e in evs:
            if e.get("event") == "battle_start":
                last_start = e
            elif e.get("event") == "battle_end":
                s = last_start or {}
                out.append({
                    "session": session, "run": run,
                    "layer": e.get("layer", ""), "node": e.get("node", ""),
                    "enc": e.get("enc", s.get("enc", "")),
                    "cat": s.get("cat", ""), "tier": s.get("tier", ""),
                    "count": s.get("count", ""), "foes": s.get("foes", ""),
                    "foes_ja": join_ja(s.get("foes", ""), enemy_ja),
                    "result": e.get("result", ""), "turns": e.get("turns", ""),
                    "hpStart": e.get("hpStart", ""), "hpLeft": e.get("hpLeft", ""),
                    "dmg_taken": int(max(0, num(e.get("hpStart")) - num(e.get("hpLeft")))),
                })
    return out


def simple_rows(events, etype, cols):
    out = []
    for e in events:
        if e.get("event") != etype:
            continue
        row = {"session": e.get("session", ""), "run": e.get("run", ""),
               "layer": e.get("layer", ""), "node": e.get("node", ""),
               "scene": e.get("scene", ""), "t": e.get("t", "")}
        for c in cols:
            row[c] = e.get(c, "")
        out.append(row)
    return out


def card_choice_rows(events):
    out = []
    for e in events:
        if e.get("event") != "card_pick":
            continue
        out.append({
            "session": e.get("session", ""), "run": e.get("run", ""),
            "node": e.get("node", ""), "t": e.get("t", ""),
            "chosen": e.get("chosen", ""), "chosen_ja": card_ja(e.get("chosen", "")),
            "offered": e.get("offered", ""), "offered_ja": join_ja(e.get("offered", ""), card_ja),
        })
    return out


def encounter_rows(script_dir):
    """encounters.json から「エンカごとの敵一覧（日本語）」の参照シートを作る。"""
    for dd in ["GameEngine/Assets/Data", "Assets/Data",
               os.path.join(script_dir, "..", "GameEngine", "Assets", "Data")]:
        arr = _load_json(os.path.join(dd, "encounters.json"), "encounters")
        if arr:
            break
    else:
        return []
    out = []
    for enc in arr:
        foes = [x.get("id", "") for x in enc.get("enemies", [])]
        out.append({
            "enc": enc.get("id", ""), "layer": enc.get("layer", ""),
            "tier": enc.get("tier", ""), "category": enc.get("category", ""),
            "weight": enc.get("weight", ""), "count": len(foes),
            "foes": ",".join(foes),
            "foes_ja": ",".join(enemy_ja(f) for f in foes),
        })
    out.sort(key=lambda r: (num(r["layer"]), str(r["category"]), str(r["enc"])))
    return out


# ---------------------------------------------------------------- サマリ集計
def build_summary(run_rows, events):
    done = [r for r in run_rows if r["result"] in ("clear", "gameover")]
    total = len(done)
    clears = sum(1 for r in done if r["result"] == "clear")
    losses = total - clears

    def avg(vals):
        vals = [num(v) for v in vals if v != "" and v is not None]
        return round(sum(vals) / len(vals), 2) if vals else 0

    secs = [r["sec"] for r in done if num(r["sec"]) > 0]     # 新規ランのみ（コンティニューはsec=0）
    stats = [
        ("総ラン数（決着のみ）", total),
        ("クリア", clears),
        ("敗北", losses),
        ("勝率(クリア率)", f"{(clears/total*100):.1f}%" if total else "-"),
        ("平均到達層", avg(r["reached_layer"] for r in done)),
        ("平均残歩数(end_node)", avg(r["end_node(steps)"] for r in done)),
        ("平均プレイ秒(新規ランのみ)", avg(secs)),
        ("平均最終デッキ枚数", avg(r["final_deckSize"] for r in done)),
        ("平均戦闘数/ラン", avg(r["battles"] for r in done)),
        ("平均被ダメ/ラン", avg(r["dmg_taken"] for r in done)),
    ]

    # 死因ランキング
    cause = Counter(r["cause"] for r in done if r["result"] == "gameover" and r["cause"])
    # カード採用率（提示 vs 選択）
    offered = Counter()
    picked = Counter()
    for e in events:
        if e.get("event") != "card_pick":
            continue
        for cid in str(e.get("offered", "")).split(","):
            cid = cid.strip()
            if cid:
                offered[cid] += 1
        ch = e.get("chosen", "")
        if ch and ch != "(skip)":
            picked[ch] += 1
    pick_rows = []
    for cid, off in offered.items():
        pk = picked.get(cid, 0)
        pick_rows.append((cid, card_ja(cid), off, pk, f"{(pk/off*100):.0f}%" if off else "-"))
    pick_rows.sort(key=lambda x: (-x[2], x[0]))       # 提示回数の多い順

    return stats, cause.most_common(), pick_rows


def relic_summary_rows(events):
    """出現したレリックを 名前・レア度・取得回数 で集計。"""
    cnt = Counter()
    for e in events:
        if e.get("event") == "relic_get":
            cnt[e.get("id", "")] += 1
        elif e.get("event") == "shop_buy" and e.get("kind") == "relic":
            cnt[e.get("id", "")] += 1
    rows = []
    for rid, c in cnt.items():
        if not rid:
            continue
        rows.append({"id": rid, "name": relic_ja(rid),
                     "rarity": RELIC_RARITY.get(rid, ""), "get_count": c})
    rows.sort(key=lambda r: (-r["get_count"], r["id"]))
    return rows


def card_coverage_rows(events):
    """全報酬対象カードを 出現/未出現ともに レア度・種類・提示/選択回数付きで一覧。"""
    offered, picked = Counter(), Counter()
    for e in events:
        if e.get("event") != "card_pick":
            continue
        for cid in str(e.get("offered", "")).split(","):
            cid = cid.strip()
            if cid:
                offered[cid] += 1
        ch = e.get("chosen", "")
        if ch and ch != "(skip)":
            picked[ch] += 1
    rows = []
    for cid in CARD_ELIGIBLE:
        off = offered.get(cid, 0)
        pk = picked.get(cid, 0)
        rows.append({
            "id": cid, "name": CARD_NAME.get(cid, cid),
            "rarity": CARD_RARITY.get(cid, ""), "type": CARD_TYPE.get(cid, ""),
            "offered_ct": off, "picked_ct": pk,
            "pick_rate": (f"{(pk/off*100):.0f}%" if off else "-"),
            "appeared": "○" if off else "×",
        })
    # 未出現(×)を上に、その中はレア度→id。出現分は提示回数が多い順
    rarity_order = {"Rare": 0, "Uncommon": 1, "Common": 2}
    rows.sort(key=lambda r: (r["appeared"] != "×", -r["offered_ct"],
                             rarity_order.get(r["rarity"], 9), r["id"]))
    return rows


# ---------------------------------------------------------------- Excel書き出し
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_table(ws, rows, headers, start_row=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=HEADER_JA.get(h, h))   # 見出しは日本語表示
    style_header(ws, len(headers), start_row)
    for i, r in enumerate(rows, start_row + 1):
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=i, column=j, value=r.get(h, ""))        # データキーは英語
            cell.font = CELL_FONT
            cell.border = BORDER
    # 列幅（表示見出しの長さ基準）
    for j, h in enumerate(headers, 1):
        label = HEADER_JA.get(h, h)
        width = max(len(str(label)), *(len(str(r.get(h, ""))) for r in rows)) if rows else len(str(label))
        ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 8), 60)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def add_sheet(wb, name, rows, headers):
    ws = wb.create_sheet(name)
    write_table(ws, rows, headers)
    if not rows:
        ws.cell(row=2, column=1, value="(データなし)").font = CELL_FONT
    return ws


def write_summary_sheet(wb, stats, cause_rank, pick_rows, files):
    ws = wb.create_sheet("summary")
    ws["A1"] = "テレメトリ集計サマリ（全体の健康診断）"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"入力ログ {len(files)} 本を集計"
    ws["A2"].font = CELL_FONT

    r = 4
    ws.cell(row=r, column=1, value="◆ 全体").font = SUB_FONT
    r += 1
    for label, val in stats:
        ws.cell(row=r, column=1, value=label).font = CELL_FONT
        c = ws.cell(row=r, column=2, value=val); c.font = Font(name=FONT, bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="◆ 死因ランキング（gameover）").font = SUB_FONT
    r += 1
    for j, h in enumerate(["敵/エンカ(id)", "敵名", "回数"]):
        c = ws.cell(row=r, column=1 + j, value=h)
        c.font = HEAD_FONT; c.fill = HEAD_FILL
    r += 1
    for cid, cnt in cause_rank:
        ws.cell(row=r, column=1, value=cid).font = CELL_FONT
        ws.cell(row=r, column=2, value=enemy_ja(cid)).font = CELL_FONT
        ws.cell(row=r, column=3, value=cnt).font = CELL_FONT
        r += 1
    if not cause_rank:
        ws.cell(row=r, column=1, value="(なし)").font = CELL_FONT
        r += 1

    # カード採用率は右側に置く
    r2 = 4
    col = 4
    ws.cell(row=r2, column=col, value="◆ カード採用率（提示 vs 選択）").font = SUB_FONT
    r2 += 1
    for j, h in enumerate(["カードid", "カード名", "提示", "選択", "採用率"]):
        c = ws.cell(row=r2, column=col + j, value=h)
        c.font = HEAD_FONT; c.fill = HEAD_FILL
    r2 += 1
    for cid, name, off, pk, rate in pick_rows:
        ws.cell(row=r2, column=col + 0, value=cid).font = CELL_FONT
        ws.cell(row=r2, column=col + 1, value=name).font = CELL_FONT
        ws.cell(row=r2, column=col + 2, value=off).font = CELL_FONT
        ws.cell(row=r2, column=col + 3, value=pk).font = CELL_FONT
        ws.cell(row=r2, column=col + 4, value=rate).font = CELL_FONT
        r2 += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    for cl in ("F", "G", "H"):
        ws.column_dimensions[cl].width = 9
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dirs", nargs="*", help="ログのディレクトリ（省略時は自動探索）")
    ap.add_argument("-o", "--out", default="telemetry_report.xlsx", help="出力xlsx")
    args = ap.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    load_name_maps(script_dir)

    dirs = args.dirs or default_dirs()
    if not dirs:
        print("ログのディレクトリが見つかりません。logs/ を指定してください。")
        sys.exit(1)

    events, files = load_events(dirs)
    if not events:
        print(f"イベントが0件でした（探索: {dirs}）")
        sys.exit(1)

    runs = group_runs(events)
    run_rows = build_run_rows(runs)
    stats, cause_rank, pick_rows = build_summary(run_rows, events)

    wb = Workbook()
    wb.remove(wb.active)   # 既定シート削除

    write_summary_sheet(wb, stats, cause_rank, pick_rows, files)

    add_sheet(wb, "encounters", encounter_rows(script_dir), [
        "enc", "layer", "tier", "category", "weight", "count", "foes", "foes_ja"])

    add_sheet(wb, "runs", run_rows, [
        "session", "run", "result", "reached_layer", "end_node(steps)", "hpLeft", "maxHp",
        "cause", "cause_ja", "sec", "final_deckSize", "gold", "battles", "wins", "losses",
        "cards_picked", "cards_skipped", "relics_got", "materials", "dmg_taken", "avg_turns",
        "final_deck", "final_deck_ja", "final_relics", "final_relics_ja"])

    add_sheet(wb, "combats", combats(runs), [
        "session", "run", "layer", "node", "enc", "cat", "tier", "count", "foes", "foes_ja",
        "result", "turns", "hpStart", "hpLeft", "dmg_taken"])

    add_sheet(wb, "card_choices", card_choice_rows(events),
              ["session", "run", "node", "t", "chosen", "chosen_ja", "offered", "offered_ja"])

    add_sheet(wb, "material_drops",
              simple_rows(events, "material_drop", ["id", "n", "rare"]),
              ["session", "run", "layer", "node", "t", "id", "n", "rare"])

    add_sheet(wb, "events",
              simple_rows(events, "event_choice", ["event", "choice"]),
              ["session", "run", "layer", "node", "t", "event", "choice"])

    add_sheet(wb, "shop",
              simple_rows(events, "shop_buy", ["id", "kind", "price"]),
              ["session", "run", "layer", "node", "t", "id", "kind", "price"])

    add_sheet(wb, "relics", relic_summary_rows(events),
              ["id", "name", "rarity", "get_count"])

    add_sheet(wb, "card_coverage", card_coverage_rows(events),
              ["id", "name", "rarity", "type", "offered_ct", "picked_ct", "pick_rate", "appeared"])

    # クラフト・休憩・強化・削除をまとめて「選択系」
    choice_rows = []

    def crow(e, kind, detail, extra, detail_ja, extra_ja):
        choice_rows.append({"session": e.get("session"), "run": e.get("run"),
                            "node": e.get("node"), "t": e.get("t"), "kind": kind,
                            "detail": detail, "detail_ja": detail_ja,
                            "extra": extra, "extra_ja": extra_ja})
    for e in events:
        et = e.get("event")
        if et == "craft":
            crow(e, "craft", e.get("id", ""), e.get("mats", ""), card_ja(e.get("id", "")), "")
        elif et == "rest_choice":
            crow(e, "rest", e.get("choice", ""), "", "", "")
        elif et == "card_upgrade":
            crow(e, "upgrade", e.get("before", ""), e.get("after", ""),
                 card_ja(e.get("before", "")), card_ja(e.get("after", "")))
        elif et == "card_remove":
            crow(e, "remove", e.get("id", ""), "", card_ja(e.get("id", "")), "")
    add_sheet(wb, "craft_rest_upgrade", choice_rows,
              ["session", "run", "node", "t", "kind", "detail", "detail_ja", "extra", "extra_ja"])

    add_sheet(wb, "feedback",
              simple_rows(events, "feedback", ["marker"]),
              ["session", "run", "layer", "node", "t", "marker"])

    add_sheet(wb, "hp_curve",
              simple_rows(events, "scene_enter", ["hp"]),
              ["session", "run", "layer", "node", "t", "scene", "hp"])

    wb.save(args.out)
    print(f"OK: {args.out} を出力（イベント {len(events)}件 / ラン {len(run_rows)}件 / ログ {len(files)}本）")


if __name__ == "__main__":
    main()
