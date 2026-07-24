# -*- coding: utf-8 -*-
"""cards.json / enemies.json / encounters.json -> GameContent.xlsx （閲覧用一覧）
Cards=種別/レアで色分け、Enemies=敵画像つき、Encounters=カテゴリで色分け。
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from cards_common import TYPE_FILL, RARITY_FILL

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "GameEngine", "Assets", "Data")
TEX_DIR = os.path.join(BASE, "GameEngine", "Assets", "Enemy")
OUT = os.path.join(BASE, "GameContent.xlsx")


def load(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


cards = load("cards.json")["cards"]
enemies = load("enemies.json")["enemies"]
encs = load("encounters.json")["encounters"]

wb = Workbook()

HEAD_FILL = PatternFill("solid", fgColor="2F3B52")
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Meiryo", size=11)
BODY = Font(name="Meiryo", size=10)
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_head(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def put(ws, r, values, fills=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.font = BODY
        cell.border = BORDER
        cell.alignment = WRAP
        if fills and (c - 1) in fills:
            cell.fill = PatternFill("solid", fgColor=fills[c - 1])


# ================= CARDS =================
ws = wb.active
ws.title = "Cards"
H = ["ID", "名前", "種別", "レア", "コスト", "射程", "範囲", "効果(main)", "値",
     "サブ", "値", "onHit", "値", "フラグ", "強化", "説明"]
W = [20, 14, 8, 10, 6, 6, 10, 12, 6, 10, 6, 12, 6, 14, 24, 34]
style_head(ws, H, W)

r = 2
for cd in cards:
    me = cd.get("mainEffect", {}) or {}
    se = cd.get("subEffect", {}) or {}
    oh = cd.get("onHitEffect", {}) or {}
    flags = [f for f in ("exhaust", "pierce", "dash") if cd.get(f)]
    if cd.get("selfDamage"):
        flags.append(f"self{cd['selfDamage']}")
    up = cd.get("upgrade", {}) or {}
    up_s = ", ".join(f"{k}:{v}" for k, v in up.items())
    rar = cd.get("rarity", "Common")
    fills = {2: TYPE_FILL.get(cd.get("type"), "FFFFFF"),
             3: RARITY_FILL.get(rar, "EFEFEF")}
    put(ws, r, [
        cd["id"], cd.get("name"), cd.get("type"), rar,
        cd.get("cost"), cd.get("range"), cd.get("rangeType"),
        me.get("type"), me.get("value"),
        se.get("type"), se.get("value"),
        oh.get("type"), oh.get("value"),
        " ".join(flags), up_s,
        (cd.get("description") or "").replace("\n", " / "),
    ], fills)
    r += 1

# ================= ENEMIES =================
ws2 = wb.create_sheet("Enemies")
H = ["画像", "ID", "HP", "ボス", "不動", "順番", "行動#", "予告", "範囲",
     "射程", "接近", "効果", "条件", "確率"]
W = [10, 16, 6, 6, 6, 6, 6, 22, 10, 6, 10, 26, 14, 6]
style_head(ws2, H, W)


def enemy_img_path(texture):
    if not texture:
        return None
    fname = texture.replace("enemy_", "") + ".png"
    p = os.path.join(TEX_DIR, fname)
    return p if os.path.exists(p) else None


r = 2
for en in enemies:
    acts = en.get("actions", [])
    base = [en.get("id"), en.get("hp"),
            "○" if en.get("isBoss") else "",
            "○" if en.get("immovable") else "",
            "○" if en.get("sequential") else ""]
    first = r
    for i, a in enumerate(acts):
        tg = a.get("target", {})
        rng = "必中" if tg.get("unavoidable") else tg.get("rangeType", "")
        eff_s = ", ".join(
            f'{e.get("kind","")}{e.get("value","") if e.get("value") is not None else ""}'
            + (f'({e.get("buff")})' if e.get("buff") else "")
            for e in a.get("effects", []))
        sel = a.get("select", {})
        cond = sel.get("condition", "")
        if cond:
            cond += f" {sel.get('conditionValue','')}"
        head = ["", base[0], base[1], base[2], base[3], base[4]] if i == 0 \
            else ["", "", "", "", "", ""]
        put(ws2, r, head + [i + 1, a.get("description", ""), rng,
                            tg.get("range", ""), tg.get("approach", ""),
                            eff_s, cond, sel.get("chance", "")])
        r += 1
    if not acts:
        put(ws2, r, ["", base[0], base[1], base[2], base[3], base[4], "", "", "", "", "", "", "", ""])
        r += 1
    # 敵ごとに区切りシェード＋画像
    for rr in range(first, r):
        ws2.cell(rr, 2).fill = PatternFill("solid", fgColor="EFEFEF")
    imgp = enemy_img_path(en.get("texture"))
    if imgp:
        img = XLImage(imgp)
        img.width = 52
        img.height = 52
        ws2.row_dimensions[first].height = 42
        ws2.add_image(img, f"A{first}")

# ================= ENCOUNTERS =================
ws3 = wb.create_sheet("Encounters")
H = ["層", "カテゴリ", "重み", "敵構成", "敵数"]
W = [6, 12, 6, 44, 6]
style_head(ws3, H, W)
CAT_FILL = {"normal": "D9EAD3", "elite": "FCE5CD", "boss": "F4CCCC"}

r = 2
for ec in encs:
    ens = ec.get("enemies", [])
    comp = ", ".join(f'{e["id"]}({e.get("col")},{e.get("row")})' for e in ens)
    cat = ec.get("category", "normal")
    put(ws3, r, [ec.get("layer", 1), cat, ec.get("weight"), comp, len(ens)],
        fills={1: CAT_FILL.get(cat, "FFFFFF")})
    r += 1

wb.save(OUT)
print("saved:", OUT)
print(f"cards={len(cards)} enemies={len(enemies)} encounters={len(encs)}")
