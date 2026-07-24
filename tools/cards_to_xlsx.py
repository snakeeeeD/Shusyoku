# -*- coding: utf-8 -*-
"""cards.json -> cards_edit.xlsx （編集用Excelを生成）"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cards_common import CARDS_JSON, CARDS_XLSX, COLS, card_to_row, TYPE_FILL, RARITY_FILL

with open(CARDS_JSON, encoding="utf-8") as f:
    cards = json.load(f)["cards"]

wb = Workbook()
ws = wb.active
ws.title = "Cards"

HEAD_FILL = PatternFill("solid", fgColor="2F3B52")
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Meiryo", size=10)
BODY = Font(name="Meiryo", size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# 編集する主要列は少し色を付ける
KEY_FILL = PatternFill("solid", fgColor="FFF9E6")
KEYCOLS = {"id", "name", "type", "cost", "range", "rangeType",
           "mainType", "mainValue", "description"}

for c, name in enumerate(COLS, 1):
    cell = ws.cell(1, c, name)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

for r, cd in enumerate(cards, 2):
    row = card_to_row(cd)
    for c, name in enumerate(COLS, 1):
        v = row.get(name)
        if isinstance(v, bool):
            v = "TRUE" if v else None
        cell = ws.cell(r, c, v)
        cell.font = BODY
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if name == "type" and row.get("type") in TYPE_FILL:
            cell.fill = PatternFill("solid", fgColor=TYPE_FILL[row["type"]])
        elif name == "rarity":
            rar = row.get("rarity") or "Common"
            cell.fill = PatternFill("solid", fgColor=RARITY_FILL.get(rar, "EFEFEF"))
        elif name in KEYCOLS:
            cell.fill = KEY_FILL

# 列幅
WIDTH = {"id": 20, "name": 14, "description": 34, "up_description": 20,
         "rangeType": 10, "mainType": 12, "subType": 10, "onHitType": 12,
         "mainBuffType": 12, "up_rangeType": 10, "tags": 10}
for c, name in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(c)].width = WIDTH.get(name, 8)

ws.freeze_panes = "D2"   # id/name/type を固定
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

wb.save(CARDS_XLSX)
print("saved:", CARDS_XLSX, "cards:", len(cards))
